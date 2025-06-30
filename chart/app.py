import openpyxl as xl
from openpyxl.chart import BarChart, Reference

print('hello folly')
def process_wookbook(filename='Book1.xlsx'):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    cell = sheet['A1']
    alt_cell = sheet.cell(row=1, column=1)

    print("sheet['A1'] =", cell.value)
    print("sheet.cell(1,1) =", alt_cell.value)
    print("Max columns:", sheet.max_column)
    print("Max rows:", sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        print(f"Row {row}")
        cell = sheet.cell(row=row, column=3)
        print("Original price:", cell.value)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row=row, column=4)
        corrected_price_cell.value = corrected_price


    values = Reference(
        sheet,
        min_row=2,
        max_row=sheet.max_row,
        min_col=4,
        max_col=4
    )
    chart = BarChart()
    chart.add_data(values, titles_from_data=False)

    sheet.add_chart(chart, "E2")


    wb.save(filename)
